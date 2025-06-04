from flask import Flask, redirect, render_template, request, jsonify, url_for, send_from_directory, send_file, session, flash
from datetime import datetime
from scripts.consultas_personalizadas import *
from scripts.iapepi import processar_arquivos
from database_utils import *
from functools import wraps
from django.shortcuts import render
from django.http import JsonResponse
from scripts.calculos import (dias_trabalhados_13, dias_trabalhados_ferias,
                              calcular_13_proporcional, ferias_proporcionais,terco_ferias, 
                              indenizacao_total, ferias_gozadas, dias_ferias_nao_gozadas)
import pandas as pd
import jinja2
import numpy as np
#from dotenv import load_dotenv
import logging
import os
import json
import subprocess
import uuid
import sqlite3
import oracledb
import pandas as pd
import openpyxl
import psutil

#load_dotenv()

app = Flask(__name__)
# Configurações de diretórios
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODELOS_DIR = os.path.join(BASE_DIR, 'modelos')
PROGRESSO_DIR = os.path.join(BASE_DIR, 'progressos')
SCRIPTS_DIR = os.path.join(BASE_DIR, 'scripts')
UPLOADS_DIR = os.path.join(BASE_DIR, 'uploads')
LOGS_DIR = os.path.join(BASE_DIR, 'logs')
ARQUIVO_EXCEL = os.path.join(BASE_DIR, 'controle_processos.xlsx')
DB_USUARIOS = os.path.join(BASE_DIR, 'usuarios.db')
DB_HISTORICO = os.path.join(BASE_DIR, 'historico.db')

# Lista de usernames permitidos
ALLOWED_USERNAMES = []
PASTA_SAIDA = os.path.join(BASE_DIR, 'arquivos_gerados')

app.config['ENV'] = os.getenv('FLASK_ENV')

os.makedirs(PASTA_SAIDA, exist_ok=True)

# ---------- LOGIN -------------- LOGIN ----------- LOGIN ----------------

app.secret_key = 'sua_chave_secreta'  

def criar_tabela():
    conn = sqlite3.connect('historico.db')  # Conectar ao banco correto
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            matricula TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            senha TEXT NOT NULL,
            permissao TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

criar_tabela() 

@app.route('/')
def index():
    if 'usuario' in session:
        return redirect(url_for('controle_processo'))  # Redireciona usuário autenticado
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    email = request.form['email']
    senha = request.form['password']

    conn = sqlite3.connect(DB_USUARIOS)
    cursor = conn.cursor()
    
    # Busca o usuário no banco
    cursor.execute("SELECT id, nome, permissao FROM usuario WHERE email = ? AND senha = ?", (email, senha))
    usuario = cursor.fetchone()
    
    conn.close()

    if usuario:
        session['usuario_id'] = usuario[0]  
        session['nome'] = usuario[1] 
        session['permissao'] = usuario[2]
        #flash('Login realizado com sucesso!', 'success')

        print(f"Usuário logado: {session}")

        return redirect(url_for('controle_processo')) 
    else:
        flash('Credenciais inválidas. Tente novamente.', 'error')
        return redirect(url_for('index'))

    
@app.route('/criar_conta')
def criar_conta():
    return render_template('criar_conta.html')

def criar_tabela_controle_processo():
    conn = sqlite3.connect('controle.db')  # Nome do banco de dados local
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS controle_processo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            periodo TEXT NOT NULL,
            numero_processo TEXT UNIQUE NOT NULL,
            orgaos TEXT NOT NULL,
            tipos TEXT NOT NULL,
            teor TEXT NOT NULL,
            numero_servidor INTEGER NOT NULL,
            valor REAL NOT NULL,
            diferenca REAL NOT NULL,
            exclusao REAL NOT NULL,
            total REAL NOT NULL,
            data_recebido TEXT,
            responsavel TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

# Executa a criação da tabela
criar_tabela_controle_processo()

@app.route('/salvar_usuario', methods=['POST'])
def salvar_usuario():
    nome = request.form['nome']
    matricula = request.form['matricula']
    email = request.form['email']
    senha = request.form['password']
    
    try:
        conn = sqlite3.connect(DB_USUARIOS)
        cursor = conn.cursor()

        # Insere os dados do novo usuário
        cursor.execute("INSERT INTO usuario (nome, matricula, email, senha, permissao) VALUES (?, ?, ?, ?, 'basico')",
                       (nome, matricula, email, senha))
        conn.commit()
        conn.close()

        flash('Cadastro realizado com sucesso! Faça login.', 'success')
        return redirect(url_for('index')) 

    except sqlite3.IntegrityError:
        flash('Erro: Email já cadastrado!', 'error')
        return redirect(url_for('criar_conta'))

def verificar_permissao(permissoes_autorizadas):
    def wrapper(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            print("Verificando permissão, sessão atual:", session)

            if 'permissao' not in session:
                print("Permissão não encontrada na sessão!")
                flash("Acesso negado!", "danger")
                return redirect(url_for('login'))

            if session['permissao'] not in permissoes_autorizadas:
                print("Permissão inválida:", session['permissao'])
                flash("Acesso negado!", "danger")
                return redirect(url_for('login'))

            print("Permissão concedida:", session['permissao'])
            return f(*args, **kwargs)
        return decorated_function
    return wrapper

@app.route('/alterar_senha', methods=['GET', 'POST'])
def alterar_senha():
    if request.method == 'POST':
        email = request.form.get('email')
        nova_senha = request.form.get('nova_senha')
        confirmar_senha = request.form.get('confirmar_senha') 

        if not email or not nova_senha or not confirmar_senha:
            flash("Preencha todos os campos!", "error")
            return render_template('alterar_senha.html')

        if nova_senha != confirmar_senha:
            flash("As senhas não coincidem!", "error")
            return render_template('alterar_senha.html')

        conn = sqlite3.connect(DB_USUARIOS)
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM usuario WHERE email = ?", (email,))
        usuario = cursor.fetchone()

        if not usuario:
            flash("E-mail não encontrado!", "error")
            return render_template('alterar_senha.html')

        # Atualizar senha no banco
        cursor.execute("UPDATE usuario SET senha = ? WHERE email = ?", (nova_senha, email))
        conn.commit()
        conn.close()

        flash("Senha alterada com sucesso! Faça login.", "success")

    return render_template('alterar_senha.html')

@app.route('/logout', methods=['POST'])
def logout():
    """Finaliza a sessão do usuário e redireciona para a página de login."""
    print(f"Finalizando sessão: {session}") 
    session.clear()  
    flash("Logout realizado com sucesso!", "success")
    return redirect(url_for('index'))  

@app.route('/pagina_restrita')
def pagina_restrita():
    if 'usuario' not in session:
        flash('Acesso negado!', 'error')  # Mensagem com categoria 'error'
        return redirect(url_for('login'))
    return render_template('pagina_restrita.html')



def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:  
            flash('Acesso restrito. Faça login para continuar.', 'error')
            return render_template('pagina_restrita.html')  
        return f(*args, **kwargs)
    return decorated_function

# ------- HISTÓRICO ------------- HISTÓRICO ---------- HISTÓRICO --------------

processes = {}

def init_db():
    conn = sqlite3.connect(DB_HISTORICO)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS historico (
            id INTEGER PRIMARY KEY,
            username TEXT,
            operation_type TEXT,
            date_time TEXT,
            file_name TEXT,
            log_file_name TEXT,
            task_id TEXT
        )
    ''')
    conn.commit()
    conn.close()

def ler_historico_db():
    conn = sqlite3.connect(DB_HISTORICO)
    c = conn.cursor()
    c.execute('SELECT * FROM historico ORDER BY date_time DESC')
    rows = c.fetchall()
    conn.close()
    historico = []
    for row in rows:
        historico.append({
            'id': row[0],
            'username': row[1],
            'operation_type': row[2],
            'date_time': row[3],
            'file_name': row[4],
            'log_file_name': row[5],
            'task_id': row[6]
        })
    return historico

def salvar_historico_db(novo_historico):
    conn = sqlite3.connect(DB_HISTORICO)
    c = conn.cursor()
    c.execute('''
        INSERT INTO historico (username, operation_type, date_time, file_name, log_file_name, task_id)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (novo_historico['username'], novo_historico['operation_type'], novo_historico['date_time'], novo_historico['file_name'], novo_historico['log_file_name'], novo_historico['task_id']))
    conn.commit()
    conn.close()

def is_user_allowed(username):
    return username in ALLOWED_USERNAMES

def read_excel(file_path):
    df = pd.read_excel(file_path, engine='openpyxl')
    return df


# ------- ROTAS ------------- ROTAS ---------- ROTAS --------------

@app.route('/operacionalizacao', methods=['GET', 'POST'])
@login_required
def operacionalizacao():
    if request.method == 'POST':
        # Processa os dados do POST aqui
        dados = request.form  # ou request.json dependendo do envio
        # Lógica de processamento
        return jsonify({"success": True, "message": "Dados processados com sucesso"})
    
    # Se for GET, apenas renderiza a página
    return render_template('operacionalizacao.html')

@app.route('/portariaAposentadoria')
@login_required
def portariaAposentadoria():
    return render_template('portariaAposentadoria.html')

@app.route('/aposentadoria')
@login_required
def aposentadoria():
    return render_template('aposentadoria.html')

@app.route('/historico')
@login_required
def historico():
    historico_dados = ler_historico_db()
    
    # Obtendo valores únicos para os dropdowns
    tipos_operacionalizacao = list(set(item["operation_type"] for item in historico_dados))
    usuarios = list(set(item["username"] for item in historico_dados))

    return render_template(
        'historico.html',
        tipos_operacionalizacao=tipos_operacionalizacao,
        usuarios=usuarios
    )

@app.route('/calculadora')
@login_required
def calculadora():
    return render_template('calculadora.html')

@app.route('/comparativo_folha', methods=['GET', 'POST'])
@login_required
def comparativo_folha():
    
    tabelas_disponiveis = ['TODAS', 'ADAPI', 'ADH', 'AGESPISA', 'AGRESP', 'ATIPI', 'BEP', 'CBMPI', 'CCOM', 'CENDROGAS', 'CEPM',  'CGEPI', 
            'COFIR', 'COJUV', 'CPVCI', 'DEFCIVIL', 'DEFENSORIA', 'DER', 'DETRAN', 'EMATER', 'EMATERPROD', 'EMGERPI', 
            'EMGERPI_CEASA', 'EMGERPI_CIDAPI', 'EMGERPI_CODIPI', 'EMGERPI_COHAB', 'EMGERPI_COMDEPI', 'EMGERPI_COMEPI',
            'EMGERPI_ETELPI_FUNART', 'EMGERPI_PIEMTUR', 'EMGERPI_PRODEPI', 'FAPEPI','FUESPI', 'FUNDESPI','FUPIP', 'GMG', 
            'IASPI', 'IDEPI', 'IMEPI', 'IMEPIPROD', 'INTERPI', 'JUCEPI', 'METRO', 'PGEPI', 'PMPI', 'SAF', 'SASC', 
            'SEADPREVPI', 'SEAGRO', 'SECID', 'SECULT', 'SEDET', 'SEDUC', 'SEDUC_PRONATEC', 'SEFAZPI', 'SEGOV', 'SEID', 
            'SEINFRA', 'SEJUSPI','SEMAR', 'SEMINPER', 'SEPLAN', 'SESAPI', 'SETRANS', 'SETUR', 'SSPPI','SURPI', 'TVANTARES', 
            'VICEGOV', 'CDTER', 'EMGERPI_CORESA', 'EMGERPI_EMATER', 'SADA','SECESP', 'SECMULHERES', 'SEDUC_PRECATORIO', 'SERES', 
            'FUNPREV' ]

    if request.method == 'POST':
        rubricas_str = request.form.get('rubrica')  # "999997,999999"
        rubricas = [r.strip() for r in rubricas_str.split(',') if r.strip()]

        mes = request.form.get('mes')
        ano = request.form.get('ano')   
        tabela_selecionada = request.form.get('tabelas')
        consulta = request.form.get('consulta')

        tabelas = [tabela_selecionada] if tabela_selecionada != "TODAS" else [t for t in tabelas_disponiveis if t != "TODAS"]

        con, cur = conectar_banco()

        # Dicionário de mapeamento para saber quantos argumentos passar
        consultas_map = {
            "RH": lambda: consulta_rh(cur, tabelas),
            "Analitico_Contracheque": lambda: consulta_analitico(cur, int(mes), int(ano), tabelas),
            "Promocao": lambda: consulta_referencias_promocao(cur, tabelas),
            "Rubricas": lambda: consulta_rubrica(cur, int(mes), int(ano), tabelas),
            "Rubricas_por_codigo": lambda: consulta_rubrica_por_codigo(cur, rubricas, int(mes), int(ano), tabelas),
            "Pensionistas_Funprev": lambda: consulta_pensionistas_funprev(cur, int(mes), int(ano)),
            "Aposentados_Funprev": lambda: consulta_aposentados(cur, int(mes), int(ano)),
            "Estagiários_Cadastro": lambda: consulta_estagiarios_cadastro(cur, tabelas),
            "Estagiários_em_Folha": lambda: consulta_estagiarios_folha(cur, int(mes), int(ano), tabelas),
            "Base_de_servidores_para_envio_ao_Picpay": lambda: consulta_picpay(cur, int(mes), int(ano), tabelas)
        }

        if consulta in consultas_map:
            try:
                df = consultas_map[consulta]()  # Executa a função correta com os argumentos necessários
                nome_arquivo = f'consulta_{consulta}_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
                caminho = os.path.join(PASTA_SAIDA, nome_arquivo)
                df.to_excel(caminho, index=False, engine='openpyxl')
                return send_file(caminho, as_attachment=True)
            except Exception as e:
                return jsonify({'status': 'erro', 'mensagem': f'Erro ao executar a consulta: {str(e)}'})
        else:
            return jsonify({'status': 'erro', 'mensagem': f'Consulta "{consulta}" não reconhecida.'})

           
    return render_template('comparativo_folha.html', tabelas=tabelas_disponiveis, colunas_consultas=COLUNAS_CONSULTAS,campos_visiveis=CAMPOS_VISIVEIS)

'''@app.route('/get_tabelas', methods=['GET'])
@login_required
def get_tabelas():
    tabelas_disponiveis = [
        'ADAPI', 'ADH', 'AGESPISA', 'AGRESP', 'ATIPI', 'BEP', 'CBMPI', 'CCOM',
        'CENDROGAS', 'CEPM', 'CGEPI', 'COFIR', 'COJUV', 'CPVCI', 'DEFCIVIL',
        'DEFENSORIA', 'DER', 'DETRAN', 'EMATER', 'EMATERPROD', 'EMGERPI',
        'EMGERPI_CEASA', 'EMGERPI_CIDAPI', 'EMGERPI_CODIPI', 'EMGERPI_COHAB',
        'EMGERPI_COMDEPI', 'EMGERPI_COMEPI', 'EMGERPI_ETELPI_FUNART',
        'EMGERPI_PIEMTUR', 'EMGERPI_PRODEPI', 'FAPEPI', 'FUESPI', 'FUNDESPI',
        'FUPIP', 'GMG', 'IASPI', 'IDEPI', 'IMEPI', 'IMEPIPROD', 'INTERPI',
        'JUCEPI', 'METRO', 'PGEPI', 'PMPI', 'SAF', 'SASC', 'SEADPREVPI',
        'SEAGRO', 'SECID', 'SECULT', 'SEDET', 'SEDUC', 'SEDUC_PRONATEC',
        'SEFAZPI', 'SEGOV', 'SEID', 'SEINFRA', 'SEJUSPI', 'SEMAR',
        'SEMINPER', 'SEPLAN', 'SESAPI', 'SETRANS', 'SETUR', 'SSPPI', 'SURPI',
        'TVANTARES', 'VICEGOV', 'CDTER', 'EMGERPI_CORESA', 'EMGERPI_EMATER',
        'SADA', 'SECESP', 'SECMULHERES', 'SEDUC_PRECATORIO', 'SERES', 'FUNPREV'
    ]
    return jsonify(tabelas_disponiveis)
'''
   
@app.route('/iapepi', methods=['GET', 'POST'])
@login_required
def iapepi():
    if request.method == 'POST':
        # Verificar se os arquivos foram enviados
        arquivo1 = request.files.get('arquivo1')
        arquivo2 = request.files.get('arquivo2')

        if not arquivo1 or not arquivo2:
            return "Por favor, envie ambos os arquivos.", 400

        # Criar o diretório 'temp' se ele não existir
        temp_dir = 'temp'
        os.makedirs(temp_dir, exist_ok=True)

        # Salvar os arquivos em um diretório temporário
        caminho_arquivo1 = os.path.join(temp_dir, arquivo1.filename)
        caminho_arquivo2 = os.path.join(temp_dir, arquivo2.filename)
        arquivo1.save(caminho_arquivo1)
        arquivo2.save(caminho_arquivo2)

        try:
            # Processar os arquivos
            nome_saida = processar_arquivos(caminho_arquivo1, caminho_arquivo2)
            caminho_saida = os.path.join(os.getcwd(), nome_saida)

            # Retornar o arquivo gerado para download
            return send_file(caminho_saida, as_attachment=True, download_name=nome_saida)
        except ValueError as e:
            return str(e), 400
        finally:
            # Remover os arquivos temporários
            if os.path.exists(caminho_arquivo1):
                os.remove(caminho_arquivo1)
            if os.path.exists(caminho_arquivo2):
                os.remove(caminho_arquivo2)

    return render_template('iapepi.html')

@app.route('/calcular', methods=['POST'])
def calcular():
    data = request.get_json()
    nome_calc = data.get('nome_calc', '')
    matricula_calc = data.get('matricula_calc', '')
    orgaos_calc = data.get('orgaos_calc', '')
    cpf_calc = data.get('cpf_calc', '')
    tipo = data['tipo']
    base_calc = float(data['base_calc'])
    cet_calc = float(data.get('cet_calc', 0.0))
    data_admissao = data['admissao']
    data_termino = data['termino']
    dias_gozados = data.get('dias_gozados', 0)

    incluir_restituicao = data.get('incluir_restituicao', False)
    liquido_calc = float(data.get('liquido_calc', 0.0))
    dias_restituir = int(data.get('dias_restituir', 0))

    dias_13 = dias_trabalhados_13(data_admissao, data_termino)
    valor_13 = calcular_13_proporcional(base_calc, dias_13)


    if tipo == "Prestador de Serviço/Temporario":
        total = valor_13
        resposta = {
            'nome_calc': nome_calc, 'matricula_calc': matricula_calc,
            'orgaos_calc': orgaos_calc, 'cpf_calc': cpf_calc,
            'tipo': tipo, 'admissao': data_admissao, 'termino': data_termino,
            'dias_13': dias_13, 'valor_13': valor_13,
            'dias_ferias': "0 dias",
            'valor_ferias_proporcionais': 0.0,
            'dias_ferias_nao_gozadas': "0 dias",
            'valor_ferias_nao_gozadas': 0.0,
            'valor_terco_ferias': 0.0,
            'total': round(total, 2)
        }

        if incluir_restituicao:
            valor_restituicao = round((liquido_calc / 30) * dias_restituir, 2)
            resposta['valor_restituicao'] = valor_restituicao
            resposta['total'] -= valor_restituicao

        return jsonify(resposta)

    dias_ferias = dias_trabalhados_ferias(data_admissao, data_termino)
    valor_ferias_prop = ferias_proporcionais(base_calc, dias_ferias)
    valor_terco = terco_ferias(valor_ferias_prop)
    dias_nao_gozadas = dias_ferias_nao_gozadas(data_admissao, data_termino, dias_gozados)
    valor_ferias_nao_gozadas = ferias_gozadas(base_calc, cet_calc, dias_ferias, dias_gozados)
    total = max(0, indenizacao_total(valor_13, valor_ferias_prop, valor_ferias_nao_gozadas, valor_terco))

    valor_restituicao = round((liquido_calc / 30) * dias_restituir, 2) if incluir_restituicao else 0.0
    total -= valor_restituicao

    return jsonify({
        'nome_calc': nome_calc, 'matricula_calc': matricula_calc,
        'orgaos_calc': orgaos_calc, 'cpf_calc': cpf_calc,
        'tipo': tipo, 'admissao': data_admissao, 'termino': data_termino,
        'dias_13': dias_13, 'valor_13': valor_13,
        'valor_ferias_proporcionais': valor_ferias_prop,
        'dias_ferias': dias_ferias,
        'valor_terco_ferias': valor_terco,
        'dias_ferias_nao_gozadas': dias_nao_gozadas,
        'valor_ferias_nao_gozadas': valor_ferias_nao_gozadas,
        'valor_restituicao': valor_restituicao,
        'total': round(total, 2)
    })


@app.route('/indenizacao')
@login_required
def indenizacao(): #login ao banco de dados
    dados = pd.read_excel('calculos_rescisao.xlsx')

    dados['Data_Admissão'] = dados['Data_Admissão'].dt.strftime('%d/%m/%Y')
    dados['Data_Desligamento'] = dados['Data_Desligamento'].dt.strftime('%d/%m/%Y')

    return render_template('indenizacao.html', dados=dados.to_dict(orient='records'))

@app.route('/salvar-pagos', methods=['POST'])
def salvar_pagos():
    try:
        dados = request.get_json()
        df = pd.read_excel('calculos_rescisao.xlsx')
        for item in dados:
            # Convertendo matrícula para string, caso ela esteja em um formato numérico no DataFrame
            df.loc[df['Matricula'].astype(str) == str(item['matricula']), 'Pago'] = item['pago']
        df.to_excel('calculos_rescisao.xlsx', index=False)
        return jsonify({'status': 'success', 'message': 'Dados salvos com sucesso!'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

def datetimeformat(value, format='%d-%m-%Y'):
    return value.strftime(format)

app.jinja_env.filters['datetimeformat'] = datetimeformat

@app.route('/buscar', methods=['POST'])
@login_required
def buscar():
    matricula = request.json.get('matricula')
    file_path = 'homologação/modelos/teste.csv'

    if not os.path.exists(file_path):
        return jsonify({"success": False, "message": "Arquivo CSV não encontrado."})

    try:
        df = pd.read_csv(file_path, sep=';', encoding='utf-8-sig')
        df['Matricula'] = df['Matricula'].astype(str).str.strip()

        resultado = df[df['Matricula'] == matricula].to_dict(orient='records')

        if resultado:
            return jsonify({"success": True, "data": resultado})
        else:
            return jsonify({"success": False, "message": "Matrícula não encontrada."})
    except Exception as e:
        return jsonify({"success": False, "message": f"Erro ao ler o arquivo CSV: {e}"})

@app.route('/salvar', methods=['POST'])
@login_required
def salvar():
    try:
        dados = request.json.get('dados')
        file_path = 'C:/Github/App_Program-1/modelos/teste.csv'

        if not os.path.exists(file_path):
            return jsonify({"success": False, "message": "Arquivo CSV não encontrado."})

        df = pd.read_csv(file_path, sep=';', encoding='utf-8-sig') 

        for item in dados:
            if 'Matricula' in item and 'Pago' in item:
                df.loc[df['Matricula'] == item['Matricula'], 'Pago'] = item['Pago']


        df.to_csv(file_path, index=False, sep=';', encoding='utf-8-sig')

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": f"Erro ao salvar dados no CSV: {e}"})

@app.route('/historico/data')
@login_required
def historico_data():
    historico_dados = ler_historico_db()
    return jsonify(historico_dados)

@app.route('/controle_processo')
@login_required
def controle_processo():
    print("Sessão atual:", session)  # Depuração: Verifica se a sessão tem as permissões certas
    return render_template('controle_processo.html')

@app.route('/criar_controle_processo')
@login_required
def criar_controle_processo():
    return render_template('criar_controle_processo.html')

@app.route('/editar_controle_processo')
@login_required
def editar_controle_processo():
    return render_template('editar_controle_processo.html')

def inicializar_excel():
    if not os.path.exists(ARQUIVO_EXCEL):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Período', 'Número do Processo', 'Órgãos', 'Tipo', 'Teor', 
                      'Número do Servidor', 'Valor', 'Diferença', 'Total', 'Status','Data', 'Responsável'])
        workbook.save(ARQUIVO_EXCEL)

#Função para buscar responsavel no pop-up.
def buscar_processos(request):
    responsavel = request.GET.get('responsavel', None)
    if not responsavel:
        return JsonResponse({'error': 'Responsável não informado'}, status=400)
    
    try:
        df = pd.read_excel(ARQUIVO_EXCEL)
        processos_filtrados = df[df['Responsável'] == responsavel]

        if processos_filtrados.empty:
            return JsonResponse({'message': 'Nenhum processo encontrado para o responsável informado'}, status=404)

        dados_processos = processos_filtrados.to_dict('records')
        return JsonResponse({'processos': dados_processos})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
#função para salvar as informações do formulário no arquivo excel.
controles = []
@app.route('/salvar_controle_processo', methods=['POST'])
@login_required
def salvar_controle_processo():
    inicializar_excel()  
    workbook = openpyxl.load_workbook('controle_processos.xlsx')
    sheet = workbook.active
    numero_processo = request.form['numero_processo']

    for row in sheet.iter_rows(values_only=True):
        if numero_processo == str(row[1]):  
            flash('Esse processo já foi cadastrado!', 'error')
            return render_template('criar_controle_processo.html')
        
    periodo = request.form['periodo']
    numero_processo = request.form['numero_processo']
    orgaos = request.form['orgaos']
    tipos = request.form.getlist('tipo[]') 
    teor = request.form['teor']
    numero_servidor = request.form['numero_servidor']

    valor_str = request.form['valor'].replace(',', '.')
    valor = float(valor_str)

    diferenca_str = request.form['diferenca'].replace(',', '.')
    diferenca = float(diferenca_str)

    exclusao_str = request.form['exclusao'].replace(',', '.')
    exclusao = float(exclusao_str)

    total = round(valor + diferenca - exclusao, 2)

    status = request.form['status']
    responsavel = request.form['responsavel']

    #separar tipo por ||
    tipo_unificado = ' || '.join(tipos)
    print("Tipos recebidos:", tipos)

    data_recebido = ''
    if status == 'Recebido':
        data_recebido = datetime.now().strftime('%d/%m/%Y - %H:%M:%S')


    controle = {
        'periodo': periodo,
        'numero_processo': numero_processo,
        'orgaos': orgaos,
        'tipos': tipo_unificado,
        'teor': teor,
        'numero_servidor': numero_servidor,
        'valor': valor,
        'diferenca': diferenca,
        'exclusao': exclusao,
        'total': total, 
        'data_recebido': data_recebido,
        'responsavel': responsavel
    }

    # Salva no Excel
    workbook = openpyxl.load_workbook('controle_processos.xlsx')
    sheet = workbook.active
    sheet.append([
        periodo, numero_processo, orgaos, tipo_unificado, teor, numero_servidor,
        valor, diferenca, exclusao, total, data_recebido, responsavel ])
    workbook.save('controle_processos.xlsx')

    flash('Processo cadastrado com sucesso!', 'success')

    controles = []  
    controles.append(controle)

    return render_template('visualizar_controle_processo.html', controles=controles)

#função para exibir as informações de arquivo excel na página de visualização.
@app.route('/visualizar_controle_processo')
@login_required
def visualizar_controle_processo():
    inicializar_excel()
    
    workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
    sheet = workbook.active
    controles = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        controles.append({
           'periodo': row[0],
            'numero_processo': str(row[1]),  
            'orgaos': row[2],
            'tipos': row[3],  
            'teor': row[4],
            'numero_servidor': row[5],
            'valor': row[6],
            'diferenca': row[7],
            'exclusao': row[8],  
            'total': row[9],  
            'data_recebido': row[10],
            'responsavel': row[11]  
        })

    responsavel_filtro = request.args.get('responsavel', '').strip()
    processo_filtro = request.args.get('processo', '').strip()
    mes_filtro = request.args.get('periodo_mes', '').strip()  
    ano_filtro = request.args.get('periodo_ano', '').strip()  

    print(f"Filtro aplicado - Mês: {mes_filtro}, Ano: {ano_filtro}")  # Debug

    # Aplicando os filtros
    if responsavel_filtro:
        controles = [c for c in controles if c['responsavel'] == responsavel_filtro]
    
    if processo_filtro:
        controles = [c for c in controles if processo_filtro.lower() in c['numero_processo'].lower()]

    if mes_filtro:
        controles = [c for c in controles if c['periodo'].startswith(mes_filtro)]  # Filtra pelo mês

    if ano_filtro:
        controles = [c for c in controles if c['periodo'].endswith(ano_filtro)]  # Filtra pelo ano

    # Resetando os filtros
    filtros_aplicados = {'responsavel': '', 'processo': '', 'periodo_mes': '', 'periodo_ano': ''}

    # Exibindo mensagem caso não haja resultados
    mensagem = ""
    if not controles:
        mensagem = "NÃO HÁ NENHUMA INFORMAÇÃO"
    
    return render_template('visualizar_controle_processo.html', controles=controles, filtros_aplicados=filtros_aplicados, mensagem=mensagem)


#Função para edição dos processos, filtrados por responsável.
@app.route('/editar_processo', methods=['GET'])
@login_required
def editar_processo():
    numero_processo = request.args.get('numero_processo')
    df = pd.read_excel(ARQUIVO_EXCEL)
    processo = df[df['Número_Processo'] == numero_processo].iloc[0].to_dict()
    return render_template('editar_processo.html', processo=processo)

#Função para salvar os processos, filtrados por responsável.
@app.route('/salvar_edicao', methods=['POST'])
@login_required
def salvar_edicao():
    try:
        # Captura o número do processo do formulário
        numero_processo = request.form.get('Número_Processo')
        if not numero_processo:
            flash('Número do Processo é necessário!', 'error')
            return redirect(url_for('index'))  # Redireciona se o número do processo não estiver presente

        # Verifica se o arquivo existe
        if not os.path.exists('controle_processos.xlsx'):
            flash('Arquivo de planilha não encontrado!', 'error')
            return redirect(url_for('index'))

        # Carrega a planilha
        workbook = openpyxl.load_workbook('controle_processos.xlsx')
        sheet = workbook.active

        # Encontra a linha do processo
        row_found = False
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value == numero_processo:
                row_found = True
                # Atualiza os dados na linha encontrada
                sheet.cell(row=row, column=1, value=request.form.get('periodo'))
                sheet.cell(row=row, column=3, value=request.form.get('orgaos'))
                sheet.cell(row=row, column=4, value=request.form.get('tipo'))
                sheet.cell(row=row, column=5, value=request.form.get('teor'))
                sheet.cell(row=row, column=6, value=request.form.get('numero_servidor'))
                sheet.cell(row=row, column=7, value=request.form.get('valor'))
                sheet.cell(row=row, column=8, value=request.form.get('diferenca'))
                sheet.cell(row=row, column=9, value=request.form.get('exclusao'))
                sheet.cell(row=row, column=12, value=request.form.get('responsavel'))
                break

        if not row_found:
            flash('Processo não encontrado na planilha!', 'error')
        # Salva a planilha
        workbook.save('controle_processos.xlsx')
        flash('Processo atualizado com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao salvar a edição: {str(e)}', 'error')
        return redirect(url_for('editar_processo', numero_processo=numero_processo))

    return redirect(url_for('visualizar_controle_processo', numero_processo=numero_processo))

#Função para apagar processo.
#@app.route('/apagar_processo', methods=['POST'])
#@login_required
#def apagar_processo():
#    dados_deletados = request.form.to_dict()
#    numero_processo = dados_deletados.get('Número_Processo')
#    responsavel = dados_deletados.get('Responsável')

#   df = pd.read_excel('controle_processos.xlsx')

#    filtro = (df['Número_Processo'] == numero_processo) & (df['Responsável'] == responsavel)
#    if not df[filtro].empty:

#        df = df.drop(df[filtro].index)

#        df.to_excel('controle_processos.xlsx', index=False)

#        return redirect(f'/visualizar_responsavel_controle?responsavel={responsavel}')
#    else:
#        return "Erro: Processo ou Responsável não encontrados.", 404

#Função download de todos os processos cadastrados.
@app.route('/download_processo')
@login_required
def download_processo():
    if not os.path.exists(ARQUIVO_EXCEL):
        inicializar_excel()
    return send_file(ARQUIVO_EXCEL, as_attachment=True)

running_processes = {}

def execute_script(script_name, operation_type, log_file_name):
    global running_processes  # Certifique-se de que estamos usando a variável global

    username = request.form['username']
    password = request.form['password']
    file = request.files['file']
    file_path = os.path.join('uploads', file.filename)
    file.save(file_path)

    if not is_user_allowed(username):
        print("Usuário inválido tentou executar um script.")
        return jsonify({"message": "Usuário inválido para operar"}), 403

    task_id = str(uuid.uuid4())
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_name = f'log_{script_name.replace(".py", "")}_{timestamp}.txt'
    print(f"Tentando iniciar o script: {script_name} com task_id: {task_id}")

    try:
        process = subprocess.Popen(['python', f'scripts/{script_name}', file_path, username, password, task_id, log_file_name])
        running_processes[task_id] = process  # Salva o processo no dicionário
        print(f"Processo salvo: {running_processes}")
    except Exception as e:
        print(f"Erro ao iniciar o script: {e}")
        return jsonify({"message": "Erro ao iniciar o script", "error": str(e)}), 500

    novo_historico = {
        "username": username,
        "operation_type": operation_type,
        "date_time": timestamp,
        "file_name": file.filename,
        "log_file_name": log_file_name,
        "task_id": task_id
    }
    salvar_historico_db(novo_historico)

    return jsonify({"message": f"Script de {operation_type} iniciado com sucesso!", 
                    "task_id": task_id, 
                    "log_file_name": log_file_name})


@app.route('/stop_script/<task_id>', methods=['POST'])
def stop_script(task_id):
    global running_processes  # Certifique-se de que está acessando a variável correta

    print(f"Tentando parar o script com task_id: {task_id}")
    print(f"Processos em execução antes de parar: {running_processes}")

    if task_id in running_processes:
        try:
            process = running_processes[task_id]
            parent = psutil.Process(process.pid)
            for child in parent.children(recursive=True):
                child.terminate()
            parent.terminate()
            del running_processes[task_id]
            print(f"Script {task_id} interrompido com sucesso.")
            return jsonify({"message": "Script interrompido com sucesso!"}), 200
        except psutil.NoSuchProcess:
            print("Processo já finalizado.")
            return jsonify({"message": "Processo já finalizado."}), 404
        except Exception as e:
            print(f"Erro ao interromper o script: {e}")
            return jsonify({"message": f"Erro ao interromper o script: {str(e)}"}), 500
    else:
        print(f"Task ID {task_id} não encontrado!")
        return jsonify({"message": "Task ID não encontrado ou já finalizado."}), 404
    
    
#Função de download do processo, filtrados por responsável.
@app.route('/download_excel', methods=['GET'])
def download_excel():
    responsavel = request.args.get('responsavel')
    if not responsavel:
        return {"error": "Responsável não informado"}, 400

    try:
        df = pd.read_excel('C:/Github/App_Program/controle_processos.xlsx')

        df_filtrado = df[df['Responsável'] == responsavel]

        arquivo_nome = f'controle_processo_{responsavel}.xlsx'
        caminho_arquivo = os.path.join(PASTA_SAIDA, arquivo_nome)

        df_filtrado.to_excel(caminho_arquivo, index=False)

        return send_file(caminho_arquivo, as_attachment=True)

    except Exception as e:
        return {"error": str(e)}, 500

#ROTAS MÓDULO PESSOA
@app.route('/scripts/script_ingresso.py', methods=['POST'])
def execute_script_ingresso():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_ingresso_{timestamp}.txt'
    return execute_script('script_ingresso.py', 'Ingresso', 'log_file_name')

@app.route('/scripts/script_promocao.py', methods=['POST'])
def execute_script_promocao():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_promocao_{timestamp}.txt'
    return execute_script('script_promocao.py', 'Promoção', 'log_file_name')

@app.route('/scripts/script_lotacao.py', methods=['POST'])
def execute_script_lotacao():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_lotacao_{timestamp}.txt'
    return execute_script('script_lotacao.py', 'lotação', 'log_file_name')

@app.route('/scripts/script_inclusao_perfil.py', methods=['POST'])
def execute_script_inclusao_perfil():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_inclusao_perfil_{timestamp}.txt'
    return execute_script('script_inclusao_perfil.py', 'Inclusão Perfil', 'log_file_name')

@app.route('/scripts/script_alterar_regime.py', methods=['POST'])
def execute_script_alterar_regime():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_alterar_regime_{timestamp}.txt'
    return execute_script('script_alterar_regime.py', 'Alterar Regime', 'log_file_name')

@app.route('/scripts/script_desligamento.py', methods=['POST'])
def execute_script_desligamento():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_desligamento_{timestamp}.txt'
    return execute_script('script_desligamento.py', 'Desligamento', 'log_file_name')

#ROTAS MÓDULO SEGURANÇA
@app.route('/scripts/script_permissoes_individuais.py', methods=['POST'])
def execute_script_permissoes_individuais():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_permissoes_individuais_{timestamp}.txt'
    return execute_script('script_permissoes_individuais.py', 'permissoes_individuais', 'log_file_name')

@app.route('/scripts/script_excluir_permissoes.py', methods=['POST'])
def execute_script_exclui_permissoes():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_excluir_permissoes_{timestamp}.txt'
    return execute_script('script_excluir_permissoes.py', 'Excluir Permissões', 'log_file_name')

#ROTAS MÓDULO FOLHA
@app.route('/scripts/script_incluir_servidor.py', methods=['POST'])
def execute_script_incluir_servidor():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_incluir_servidor_{timestamp}.txt'
    return execute_script('script_incluir_servidor.py', 'incluir_servidor', 'log_file_name')

@app.route('/scripts/script_ger_ev_cal.py', methods=['POST'])
def execute_script_importar_ger_ev_cal():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_ger_ev_cal_{timestamp}.txt'
    return execute_script('script_ger_ev_cal.py', 'Importar Ger. Ev. Cal.', 'log_file_name')

#ROTAS MÓDULO TABELAS BASICAS
@app.route('/scripts/script_organograma.py', methods=['POST'])
def execute_script_organograma():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_orgonograma_{timestamp}.txt'
    return execute_script('script_organograma.py', 'organograma', 'log_file_name')

#ROTAS MÓDULO FUNPREV
@app.route('/scripts/script_ingresso_pecunia.py', methods=['POST'])
def execute_script_ingresso_pecunia():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_ingresso_pecunia_{timestamp}.txt'
    return execute_script('script_ingresso_pecunia.py', 'Ingresso Pecunia', 'log_file_name')

@app.route('/scripts/script_aposentadoria.py', methods=['POST'])
def execute_script_aposentadoria():
    #timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    #log_file_name = f'log_aposentadoria_{timestamp}.txt'
    return execute_script('script_aposentadoria.py', 'Aposentadoria', 'log_file_name')

# Define o filtro de log para suprimir solicitações da rota /progress/<task_id>
class FilterProgressRequests(logging.Filter):
    def filter(self, record):
        return "/progress/" not in record.getMessage()

# Configura o log para ignorar a rota /progress/<task_id>
log = logging.getLogger('werkzeug')
log.addFilter(FilterProgressRequests())

@app.route('/progress/<task_id>')
def get_progress(task_id): 
    progress_file_path = os.path.join(PROGRESSO_DIR, f'progress_{task_id}.json')
    if os.path.exists(progress_file_path):
        with open(progress_file_path, 'r') as progress_file:
            progress_data = json.load(progress_file)
        return jsonify(progress_data)
    else:
        return jsonify({'task_id': task_id, 'progress': 0})

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory('modelos', filename)

@app.route('/uploads/<filename>')
def download_uploaded_file(filename):
    return send_from_directory('uploads', filename)

@app.route('/logs/<filename>')
def download_log_file(filename):
    return send_from_directory('logs', filename)

@app.errorhandler(Exception)
def handle_exception(e):
    # Lida com exceções genéricas
    logging.exception("Erro inesperado: %s", str(e))
    return jsonify({"error": "Ocorreu um erro inesperado. Tente novamente mais tarde."}), 500

@app.errorhandler(404)
def handle_404(e):
    return jsonify({"error": "Rota não encontrada"}), 404

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=80, host='0.0.0.0')
